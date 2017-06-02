import openpyxl
from openpyxl.utils import get_column_letter
import matrix

class Excel:
    def __init__(self):
        self.a = matrix.Matrix([[0]], "Initial matrix")
        self.makedafault()
        self.commands = {
            "none": 0,
            "exit": 1,
            "test": 2,
            "clear": 3,
            "help": 4,
            "new": 5,
            "show slist": 6,
            "mk": 8,
            "QR": 9,
            "WR": 10

        }
        pass

    def makedafault(self):
        self.setpath("./")
        self.setfilename("file.xlsx")
        self.setsheetname("Sheet1")
        pass

    def importparam(self):
        pass


    def setpath(self, path):
        self.path = path
        pass

    def setsheetname(self, sheetname):
        self.sheetname = sheetname
        pass

    def setfilename(self, filename):
        self.filename = filename
        pass

    def getfilepath(self):
        return self.path + self.filename
        pass

    def enterCommand(self):
        command = "0"
        print('')
        print("Enter command (help for Q&A)")
        while (command not in self.commands):
            command = input("->")
            if (command not in self.commands):
                print("There is no such command")
            else:
                return self.commands[command]

    def showCommands(self):
        print('')
        print("Commands...")
        for item in self.commands:
            print(str(item) + ": " + str(self.commands[item]))

    def showHelp(self):
        print('')
        print("Help v0.001")
        print("Author of this program: Sir Oleksiy Polshchak")
        self.showCommands()

    def dostaff(self):
        task = 0
        while (task != 1):
            print('')
            print("Work with Excel")
            print('')
            task = self.enterCommand()
            if (task == 2):
                self.dostaff()
            elif (task == 3):
                pass
            elif (task == 4):
                self.showHelp()
            elif (task == 5):
                self.inputnewdata()
                pass
            elif (task == 6):
                self.a.showmatrix()
                pass
            elif (task == 8):
                self.makedafault()
            elif (task == 10):
                self.write_data()
        pass

    def inputnewdata(self):
        task = 0
        self.a = self.inputmatrix()

    def inputmatrix(self):
        print('')
        i = 0
        task = 0
        return self.transferlist(param='square')


    def transferlist(self, param):
        if (param == "int"):
            return list(map(int, input("-> ").split()))

        elif (param == "float"):
            return list(map(float, input("-> ").split()))
        #elif (param == "default"):
        #    pass
        elif (param == "square"):
            filepath = self.getfilepath()
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.get_sheet_by_name(self.sheetname)
            # sheet.cell(row=i, column=self.drow).value
            return matrix.Matrix([[sheet.cell(row=k, column=i).value for i in range(1, sheet.max_column + 1)] for k in range(1, sheet.max_column + 1)], "Initial matrix")
            pass
        else:
            return 1

    def write_data(self):
        wb = openpyxl.Workbook()
        dest_filename = 'empty_book.xlsx'
        ws1 = wb.active
        ws1.title = "range names"

        for row in range(1, 40):
            ws1.append(range(600))

        ws2 = wb.create_sheet(title="Pi")

        ws2['F5'] = 3.14

        ws3 = wb.create_sheet(title="Data")
        for row in range(10, 20):
            for col in range(27, 54):
                _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
        print(ws3['AA10'].value)
        wb.save(filename=dest_filename)